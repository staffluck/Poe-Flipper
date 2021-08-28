import argparse
import requests
import xlsxwriter
import os
import openpyxl
from json.decoder import JSONDecodeError
from typing import List

ALL_POSSIBLE_CATEGORIES = ['accessory', 'armour', 'weapon', 'jewel', 'uniqueMap']
POETRADE_HEADERS = {'User-Agent': 'agent47daun@gmail.com'}


class PoeFlipper:

    def __init__(self, league, filename="item_table.xlsx"):
        self.league = league
        self.filename = filename

        self.parsed_items_data = []
        self.result_items_data = []

        self.poewatch_get_url = "https://api.poe.watch/get?category={}&league={league}".format("{}", league=league)
        self.poetrade_search = "https://www.pathofexile.com/api/trade/search/{}".format(league)
        self.poetrade_fetch = "https://www.pathofexile.com/api/trade/fetch/{}?query={}"
        self.poetrade_stats = "https://www.pathofexile.com/api/trade/data/stats"

    def convert_items_data(self) -> None:

        # Converting items stats into data like POETRADE_ITEM_ID:ROLL_RANGE
        def convert_mod(mod: str) -> List:
            try:
                if mod.startswith("(") or mod.startswith("+("):
                    mod_raw = mod.split("(")[1].split(")")
                    mod_range = mod_raw[0]
                    mod_text = "#" + mod_raw[1]
                    if mod.startswith("+"):
                        mod_text = "+" + mod_text
                    mod_range = mod_range.split("-")
                elif mod.startswith("Socketed Gems are"):
                    mod_range = [mod[36:39].strip(), ]
                    mod_text = mod[:36] + " # " + mod[39:]
                    mod_text = mod_text.replace("  ", " ")
                elif mod.count("(") == 2:
                    # How it works:
                    # mod = Adds (65-75) to (100-110) Physical Damage
                    # mod.split("(") => ['Adds ', '65-75) to ', '100-110) Physical Damage']
                    # mod.split("(")[1].split(")") => ('65-75', ' to ')
                    # mod.split("(")[2].split(")") => ('100-110', ' Physical Damage')
                    # mod.split("(")[0] + "#" + to + "#" + bonus => Adds # to # Physical Damage
                    mod_range_first, to = mod.split("(")[1].split(")")
                    mod_range_second, bonus = mod.split("(")[2].split(")")
                    mod_range = [mod_range_first.split("-"), mod_range_second.split("-")]
                    mod_text = mod.split("(")[0] + "#" + to + "#" + bonus
                else:
                    mod_raw = mod.split("(")[1].split(")")
                    mod_range = mod_raw[0]
                    mod_text = "#" + mod_raw[1]
            except IndexError:
                #  Handling mods without range
                return [False, False]

            try:
                if converted_stats_mods.get(mod_text):
                    mod_id = converted_stats_mods[mod_text]
                elif converted_stats_mods.get(mod_text + " (Local)"):
                    mod_id = converted_stats_mods[mod_text + " (Local)"]
                else:
                    mod_id = converted_stats_mods[mod_text[1:]]
            except KeyError:
                #  Handling mods that have no POETRADE_ITEM_ID conversion
                return [False, False]

            return mod_id, mod_range

        try:
            stats_request = requests.get(self.poetrade_stats, headers=POETRADE_HEADERS)
            stats_data = stats_request.json()
        except JSONDecodeError:
            print("Stats fetch failed. Try again in 10 sec..")
            raise SystemExit

        converted_stats_mods = {}
        stats = stats_data['result']
        for stat in stats:
            for b in stat['entries']:
                converted_stats_mods[b['text']] = b['id']

        for item in self.parsed_items_data:
            links_price = False  # Price depends on links(6-links)
            explicits_converted = {}  # Price depends on explicit rolls
            implicits_converted = {}  # Price depends on implicit rolls if item corrupted

            if item.get('explicits'):
                explicits = item['explicits'].split("&")
                for explicit in explicits:
                    explicit_id, explicit_range = convert_mod(explicit)
                    if not explicit_id:
                        continue
                    explicits_converted[explicit_id] = explicit_range

            if item.get('implicits'):
                implicits = item['implicits'].split("&")
                for implicit in implicits:
                    implicit_id, implicit_range = convert_mod(implicit)
                    if not implicit_id:
                        continue
                    implicits_converted[implicit_id] = implicit_range

            if item['group'] == "bodyarmours":
                links_price = True

            self.result_items_data.append({
                "category": item['category'],
                "group": item['group'],
                "name": item['name'],
                "mean": item['mean'],
                "explicits": explicits_converted,
                "implicits": implicits_converted,
                "depends_on_links": links_price,
            })

    def parse_file(self) -> None:
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=6):
            self.parsed_items_data.append({
                "category": row[0].value,
                "group": row[1].value,
                "name": row[2].value,
                "explicits": row[3].value,
                "implicits": row[4].value,
                "mean": row[5].value
            })
        self.convert_items_data()

    # def check_price(self, item) -> None:
    #     search_query = {
    #         "query": {
    #             "status": {
    #                 "option": "online"
    #             },
    #             "name": item["name"],

    #         }
    #     }
    #     try:
    #         request = requests.post(self.poetrade_search, data=)
    #     except JSONDecodeError as e:
    #         pass

    def start(self) -> None:
        if os.path.isfile(self.filename):
            print("Found {} table. Processing..".format(self.filename))
        else:
            print("File not found. Generating table.. ")
            self.generate_items_table(['armour', 'accessory', 'weapon'])
            return self.start()

        self.parse_file()
        # for item in self.result_items_data:
        #     zxc = self.check_price(item)

    def generate_items_table(self, categories_to_flip, custom_filename=None) -> None:
        workbook = xlsxwriter.Workbook("item_table.xlsx" if not custom_filename else "{}.xlsx".format(custom_filename))
        ws = workbook.add_worksheet()

        row = 1
        for category in categories_to_flip:
            try:
                request = requests.get(self.poewatch_get_url.format(category))
                items_data = request.json()
            except JSONDecodeError as e:
                print(e, request)
                continue

            ws.write(0, 0, "Category")
            ws.write(0, 1, "Group")
            ws.write(0, 2, "Name")
            ws.write(0, 3, "Explicits")
            ws.write(0, 4, "Implicits")
            ws.write(0, 5, "Mean")

            column = 0
            for item in items_data:
                explicits = ""
                implicits = ""
                if item.get('explicits'):
                    for i in item['explicits']:
                        explicits += "{}&".format(i)
                if item.get('implicits'):
                    for i in item['implicits']:
                        implicits += "{}&".format(i)

                ws.write(row, column, item['category'])
                ws.write(row, column+1, item['group'])
                ws.write(row, column+2, item['name'])
                ws.write(row, column+3, explicits)
                ws.write(row, column+4, implicits)
                ws.write(row, column+5, item['mean'])
                row += 1
            row += 1

        workbook.close()


def init_argparse() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        usage="%(prog)s [OPTION]...",
    )
    parser.add_argument(
        "-gt", "--generate-table", nargs="+",
        help="Generate tables with selected categories. List of categories: https://api.poe.watch/categories. Default armour accessory weapon",
        type=str,
    )
    parser.add_argument("-cf", "--custom-filename",
                        help="Work only in pair with --generate-table(-gt). Provides custom filename for generated table.",
                        )
    parser.add_argument('-if', '--import-file',
                        help="Use custom excel table. File must be in the same folder as script")
    return parser


def main():
    parser = init_argparse()
    args = parser.parse_args()

    custom_file = "item_table.xlsx"
    if args.import_file:
        if not os.path.isfile(args.import_file):
            print("{} Not found".format(args.import_file))
            return 0
        custom_file = args.import_file
    flipper = PoeFlipper("Expedition", custom_file)

    if args.custom_filename and not args.generate_table:
        print("-cf works only in pair with --generate-table(-gt)")
        return 0
    if args.generate_table:
        custom_filename = "None"
        for arg in args.generate_table: 
            if arg not in ALL_POSSIBLE_CATEGORIES:
                print("{} category not supported. Skip..".format(arg))
                args.generate_table.remove(arg)
        if args.custom_filename:
            custom_filename = args.custom_filename[0]
        flipper.generate_items_table(args.generate_table, custom_filename)

    flipper.start()


if __name__ == "__main__":
    main()