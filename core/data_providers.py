from abc import ABCMeta, abstractmethod
from typing import List, Optional

import openpyxl
import xlsxwriter


class RegistryBase(ABCMeta):

    REGISTRY = {}

    def __new__(cls, cls_name, bases, attrs):
        created_class = super().__new__(cls, cls_name, bases, attrs)
        if bases:  # BaseDateProvider не должен регистрироваться
            cls.REGISTRY[attrs["file_format"]] = created_class
        return created_class

    @classmethod
    def get_registry(cls):
        return dict(cls.REGISTRY)


class BaseDataProvider(metaclass=RegistryBase):
    base_filename = "base_filename.base"

    def __init__(self, filename: Optional[str] = None):
        self.filename = filename if filename else self.base_filename

    @abstractmethod
    def generate_file(categories: List[dict]) -> None:
        pass

    @abstractmethod
    def parse_file() -> List:
        pass

class XlsxDataProvider(BaseDataProvider):
    file_format = "xlsx"
    base_filename = "item_table.xlsx"

    def generate_file(self, categories: List[dict]) -> None:
        workbook = xlsxwriter.Workbook(self.filename)
        ws = workbook.add_worksheet()
        ws.write(0, 0, "Category")
        ws.write(0, 1, "Group")
        ws.write(0, 2, "Name")
        ws.write(0, 3, "Explicits")
        ws.write(0, 4, "Implicits")
        ws.write(0, 5, "Mean")

        row = 1
        for category in categories:
            column = 0
            for item in category:
                explicits = ""
                implicits = ""
                if item.get('explicits'):
                    for i in item['explicits']:
                        explicits += "{}&".format(i)
                if item.get('implicits'):
                    for i in item['implicits']:
                        implicits += "{}&".format(i)

                ws.write(row, column, item['category'])
                ws.write(row, column + 1, item['group'])
                ws.write(row, column + 2, item['name'])
                ws.write(row, column + 3, explicits)
                ws.write(row, column + 4, implicits)
                ws.write(row, column + 5, item['mean'])
                row += 1
            row += 1

        workbook.close()

    def parse_file(self) -> List:
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active

        parsed_items_data = []
        for row in sheet.iter_rows(min_row=2, max_col=6):
            parsed_items_data.append({
                "category": row[0].value,
                "group": row[1].value,
                "name": row[2].value,
                "explicits": row[3].value,
                "implicits": row[4].value,
                "mean": row[5].value
            })

        return parsed_items_data
